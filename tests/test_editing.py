import os
import io
import subprocess
import sys
from pathlib import Path
from datetime import date
import pytest
from openpyxl import load_workbook

os.environ.setdefault("DATABASE_URL", "sqlite:///./test_editing.db")
os.environ.setdefault("SECRET_KEY", "test-editing-secret")

from fastapi.testclient import TestClient
import app as app_module
from backend import models


def reset_db():
    models.Base.metadata.drop_all(bind=models.engine)
    models.Base.metadata.create_all(bind=models.engine)
    db = models.SessionLocal()
    admin = models.User(full_name="Администратор", login="edit-admin", password_hash=app_module.pwd_hash("pass"), role=models.UserRole.ADMIN, is_active=True)
    logist = models.User(full_name="Логист", login="edit-logist", password_hash=app_module.pwd_hash("pass"), role=models.UserRole.LOGIST, is_active=True)
    driver = models.User(full_name="Водитель Старый", login="edit-driver", password_hash=app_module.pwd_hash("pass"), role=models.UserRole.DRIVER, is_active=True)
    vt = models.VehicleType(name="Пухтовоз", kind=models.TripType.PUKHTOVOZ)
    db.add_all([admin, logist, driver, vt]); db.flush()
    vehicle = models.Vehicle(name="КАМАЗ старый", plate="А111АА78", type_id=vt.id, capacity=10, is_active=True)
    customer = models.Customer(name="Заказчик старый", address="Старый адрес")
    cargo = models.CargoType(name="Груз старый", unit="м3")
    polygon = models.Polygon(name="Полигон старый", address="Старый полигон")
    db.add_all([vehicle, customer, cargo, polygon]); db.flush()
    tariff = models.Tariff(title="Тариф старый", vehicle_type_id=vt.id, kind=models.TripType.PUKHTOVOZ, trip_price=1000, is_active=True)
    db.add(tariff); db.flush()
    trip = models.TripRequest(
        number="П-EDIT", planned_date=date(2026, 8, 11), planned_time="08:00",
        driver_id=driver.id, vehicle_id=vehicle.id, customer_id=customer.id,
        cargo_type_id=cargo.id, polygon_id=polygon.id, tariff_id=tariff.id,
        load_address="Старая подача", unload_address="Старая выгрузка",
        route_name="Старый маршрут", km=10, volume=5, trips_count=1,
        kind=models.TripType.PUKHTOVOZ, status=models.RequestStatus.ASSIGNED,
        comment="Старый комментарий",
    )
    db.add(trip); db.commit()
    return db, admin, logist, driver, vt, vehicle, customer, cargo, polygon, tariff, trip


def client_as(user):
    app_module.app.dependency_overrides[app_module.get_current_user] = lambda: user
    return TestClient(app_module.app, headers={"Origin": "http://testserver"})


def test_admin_can_edit_request_and_driver_cannot():
    db, admin, _, driver, _, vehicle, customer, cargo, polygon, tariff, trip = reset_db()
    trip.comment = None
    db.commit()
    client = client_as(admin)
    page = client.get(f"/requests/{trip.id}/edit")
    assert page.status_code == 200
    assert "Редактирование заявки" in page.text
    assert 'value="П-EDIT"' in page.text
    assert '>None</textarea>' not in page.text
    response = client.post(f"/requests/{trip.id}/edit", data={
        "number": "П-EDIT-2", "planned_date": "2026-08-12", "planned_time": "09:30",
        "driver_id": str(driver.id), "vehicle_id": str(vehicle.id),
        "customer_id": str(customer.id), "cargo_type_id": str(cargo.id),
        "polygon_id": str(polygon.id), "tariff_id": str(tariff.id),
        "load_address": "Новая подача", "unload_address": "Новая выгрузка",
        "route_name": "Новый маршрут", "km": "20", "volume": "7", "trips_count": "2",
        "kind": "пухтовоз", "comment": "Новый комментарий",
    }, follow_redirects=False)
    assert response.status_code == 302
    db.expire_all()
    saved = db.query(models.TripRequest).filter_by(id=trip.id).one()
    assert saved.number == "П-EDIT-2"
    assert saved.route_name == "Новый маршрут"
    assert saved.trips_count == 2
    assert saved.status == models.RequestStatus.ASSIGNED

    driver_client = client_as(driver)
    assert driver_client.get(f"/requests/{trip.id}/edit").status_code == 403
    db.close()


def test_trip_time_contact_and_planned_tonnage_are_saved_and_visible():
    db, admin, _, driver, _, vehicle, customer, cargo, polygon, tariff, trip = reset_db()
    client = client_as(admin)
    response = client.post(f"/requests/{trip.id}/edit", data={
        "number": trip.number, "planned_date": "2026-08-13", "planned_time": "15:30",
        "driver_id": str(driver.id), "vehicle_id": str(vehicle.id),
        "customer_id": str(customer.id), "cargo_type_id": str(cargo.id),
        "polygon_id": str(polygon.id), "tariff_id": str(tariff.id),
        "load_address": "Объект клиента", "unload_address": "Полигон",
        "route_name": "Клиент — полигон", "km": "25", "volume": "12",
        "tonnage": "8.5", "trips_count": "1", "kind": "пухтовоз",
        "site_contact_name": "Иван Петров", "site_contact_phone": "+79991234567",
        "site_contact_comment": "Позвонить за 20 минут", "comment": "",
    }, follow_redirects=False)
    assert response.status_code == 302
    db.expire_all()
    saved = db.query(models.TripRequest).filter_by(id=trip.id).one()
    assert saved.planned_time == "15:30"
    assert saved.tonnage == 8.5
    assert saved.site_contact_name == "Иван Петров"
    assert saved.site_contact_phone == "+79991234567"
    assert saved.site_contact_comment == "Позвонить за 20 минут"

    detail = client.get(f"/requests/{trip.id}").text
    assert "13.08.2026, 15:30" in detail
    assert "8.5 т" in detail
    assert 'href="tel:+79991234567"' in detail
    assert "Иван Петров" in detail and "Позвонить за 20 минут" in detail
    assert "13.08.2026, 15:30" in client.get("/requests").text
    assert "13.08.2026, 15:30" in client.get("/pukhtovoz").text
    assert "13.08.2026, 15:30" in client.get("/reports").text
    driver_page = client_as(driver).get("/driver").text
    assert "13.08.2026, 15:30" in driver_page
    assert customer.name in driver_page and "Объект клиента" in driver_page
    db.close()


def test_new_trip_saves_time_contact_and_planned_tonnage():
    db, admin, _, driver, _, vehicle, customer, cargo, polygon, tariff, _ = reset_db()
    client = client_as(admin)
    response = client.post("/requests/new", data={
        "number": "П-NEW-FIELDS", "planned_date": "2026-08-14", "planned_time": "07:45",
        "driver_id": str(driver.id), "vehicle_id": str(vehicle.id),
        "customer_id": str(customer.id), "cargo_type_id": str(cargo.id),
        "polygon_id": str(polygon.id), "tariff_id": str(tariff.id),
        "load_address": "Новый объект", "unload_address": "Новый полигон",
        "route_name": "Новый маршрут", "km": "30", "volume": "16",
        "tonnage": "11.25", "trips_count": "1", "kind": "пухтовоз",
        "site_contact_name": "Олег", "site_contact_phone": "+79990001122",
        "site_contact_comment": "Въезд через КПП", "comment": "",
    }, follow_redirects=False)
    assert response.status_code == 303
    saved = db.query(models.TripRequest).filter_by(number="П-NEW-FIELDS").one()
    assert saved.planned_time == "07:45"
    assert saved.tonnage == 11.25
    assert (saved.site_contact_name, saved.site_contact_phone, saved.site_contact_comment) == (
        "Олег", "+79990001122", "Въезд через КПП",
    )
    db.close()


def test_driver_completes_trip_with_actual_tonnage_only_in_actual_fields():
    db, _, _, driver, _, _, _, _, _, _, trip = reset_db()
    trip.status = models.RequestStatus.IN_WORK
    trip.tonnage = 10
    trip.actual_tonnage = None
    db.commit()
    client = client_as(driver)

    response = client.post(f"/requests/{trip.id}/complete", data={
        "actual_km": "12", "actual_volume": "0", "actual_tonnage": "8.75",
        "comment": "Вес по талону",
    }, follow_redirects=False)
    assert response.status_code == 302
    db.expire_all()
    saved = db.query(models.TripRequest).filter_by(id=trip.id).one()
    assert saved.status == models.RequestStatus.DRIVER_COMPLETED
    assert saved.tonnage == 10
    assert saved.actual_tonnage == 8.75

    saved.status = models.RequestStatus.IN_WORK
    db.commit()
    rejected = client.post(f"/requests/{trip.id}/complete", data={
        "actual_km": "12", "actual_volume": "0", "actual_tonnage": "-1",
    }, follow_redirects=False)
    assert rejected.status_code == 400
    db.close()


def test_report_exports_include_time_and_tonnage_columns():
    db, admin, _, _, _, _, _, _, _, _, trip = reset_db()
    trip.planned_time = "15:30"
    trip.tonnage = 9.5
    trip.actual_tonnage = 8.75
    db.commit()
    client = client_as(admin)

    csv_response = client.get("/export/report.csv")
    csv_text = csv_response.content.decode("utf-8-sig")
    assert csv_response.status_code == 200
    assert "Время" in csv_text and "Тоннаж план, т" in csv_text and "Тоннаж факт, т" in csv_text
    assert "15:30" in csv_text and "9.5" in csv_text and "8.75" in csv_text

    xlsx_response = client.get("/export/report.xlsx")
    workbook = load_workbook(io.BytesIO(xlsx_response.content))
    rows = list(workbook.active.iter_rows(values_only=True))
    assert "Время" in rows[0]
    assert "Тоннаж план, т" in rows[0] and "Тоннаж факт, т" in rows[0]
    assert "15:30" in rows[1] and 9.5 in rows[1] and 8.75 in rows[1]
    db.close()


def test_driver_facts_attachments_and_access_control(tmp_path, monkeypatch):
    db, admin, logist, driver, vt, vehicle, customer, cargo, polygon, tariff, trip = reset_db()
    other_driver = models.User(
        full_name="Чужой водитель", login="other-driver", password_hash=app_module.pwd_hash("pass"),
        role=models.UserRole.DRIVER, is_active=True,
    )
    db.add(other_driver)
    trip.status = models.RequestStatus.IN_WORK
    db.commit()
    monkeypatch.setattr(app_module, "UPLOAD_DIR", str(tmp_path))

    driver_client = client_as(driver)
    response = driver_client.post(f"/requests/{trip.id}/complete", data={
        "actual_km": "15", "actual_volume": "5", "actual_tonnage": "4.2",
        "is_empty_run": "1", "empty_run_comment": "Не смог загрузиться",
        "has_downtime": "1", "downtime_minutes": "90",
        "downtime_comment": "Ждал пропуск", "comment": "Фото приложено",
    }, follow_redirects=False)
    assert response.status_code == 302
    db.refresh(trip)
    assert trip.is_empty_run is True and trip.empty_run_comment == "Не смог загрузиться"
    assert trip.has_downtime is True and trip.downtime_minutes == 90
    assert trip.downtime_comment == "Ждал пропуск"

    files = [
        ("files", (f"photo-{index}.jpg", b"\xff\xd8\xfftest", "image/jpeg"))
        for index in range(1, 6)
    ]
    uploaded = driver_client.post(f"/requests/{trip.id}/attachments", files=files, follow_redirects=False)
    assert uploaded.status_code == 302
    attachments = db.query(models.Attachment).filter_by(trip_request_id=trip.id).all()
    assert len(attachments) == 5
    assert all(Path(item.path).parent == tmp_path for item in attachments)

    overflow = driver_client.post(
        f"/requests/{trip.id}/attachments",
        files={"files": ("sixth.jpg", b"\xff\xd8\xfftest", "image/jpeg")},
    )
    assert overflow.status_code == 400

    logist_client = client_as(logist)
    detail = logist_client.get(f"/requests/{trip.id}").text
    assert "Холостой прогон" in detail and "90 мин" in detail and "photo-1.jpg" in detail
    downloaded = logist_client.get(f"/attachments/{attachments[0].id}")
    assert downloaded.status_code == 200 and downloaded.content == b"\xff\xd8\xfftest"
    Path(attachments[0].path).unlink()
    restored_from_db = logist_client.get(f"/attachments/{attachments[0].id}")
    assert restored_from_db.status_code == 200 and restored_from_db.content == b"\xff\xd8\xfftest"

    outsider = client_as(other_driver)
    assert outsider.get(f"/attachments/{attachments[0].id}").status_code == 403
    denied = outsider.post(
        f"/requests/{trip.id}/attachments",
        files={"files": ("foreign.jpg", b"\xff\xd8\xfftest", "image/jpeg")},
    )
    assert denied.status_code == 403
    db.close()


def test_driver_day_report_is_editable_unique_and_role_scoped():
    db, admin, logist, driver, vt, vehicle, customer, cargo, polygon, tariff, trip = reset_db()
    trip.planned_date = date(2026, 8, 13)
    other_driver = models.User(
        full_name="Другой водитель", login="day-other", password_hash=app_module.pwd_hash("pass"),
        role=models.UserRole.DRIVER, is_active=True,
    )
    db.add(other_driver); db.commit()
    client = client_as(driver)
    created = client.post("/driver/day-report", data={
        "report_date": "2026-08-13", "vehicle_id": str(vehicle.id),
        "total_km": "180", "odometer": "125000", "fuel_liters": "90",
        "comment": "Смена завершена",
    }, follow_redirects=False)
    assert created.status_code == 302
    report = db.query(models.DriverDayReport).filter_by(driver_id=driver.id).one()
    assert report.total_km == 180 and report.odometer == 125000 and report.fuel_liters == 90

    updated = client.post("/driver/day-report", data={
        "report_date": "2026-08-13", "vehicle_id": str(vehicle.id),
        "total_km": "195", "odometer": "125015", "fuel_liters": "95",
        "comment": "Исправлено водителем",
    }, follow_redirects=False)
    assert updated.status_code == 302
    assert db.query(models.DriverDayReport).filter_by(driver_id=driver.id).count() == 1
    db.refresh(report)
    assert report.total_km == 195 and report.comment == "Исправлено водителем"
    assert "Редактировать" in client.get("/driver/day-report").text

    logist_page = client_as(logist).get("/reports/day").text
    assert driver.full_name in logist_page and "195" in logist_page and "Исправлено водителем" in logist_page
    assert client_as(other_driver).get(f"/driver/day-report/{report.id}/edit").status_code in {403, 404}
    db.close()


def test_polygon_cost_report_uses_polygon_tariff_not_driver_payment():
    db, admin, logist, driver, vt, vehicle, customer, cargo, polygon, tariff, trip = reset_db()
    client = client_as(admin)
    saved = client.post(f"/settings/polygons/{polygon.id}/edit", data={
        "name": "Шишкино — сортировка / полигон", "address": "Шишкино",
        "contact": "Диспетчер", "phone": "+79995554433", "comment": "Приём до 20:00",
        "entry_notes": "Заезд через вторые ворота", "navigator_url": "https://yandex.ru/maps/?rtext=Шишкино",
        "calculation_method": "tonnes", "volume_rate": "0", "tonnage_rate": "750",
        "waste_types": "ТБО, грунт",
    }, follow_redirects=False)
    assert saved.status_code == 302
    db.refresh(polygon)
    assert polygon.calculation_method == "tonnes" and polygon.tonnage_rate == 750
    assert polygon.entry_notes == "Заезд через вторые ворота"
    for unsafe_url in ("javascript:alert(1)", "data:text/html,x", "https:javascript-alert"):
        rejected = client.post(f"/settings/polygons/{polygon.id}/edit", data={
            "name": polygon.name, "address": polygon.address or "", "contact": polygon.contact or "",
            "phone": polygon.phone or "", "comment": polygon.comment or "",
            "entry_notes": polygon.entry_notes or "", "navigator_url": unsafe_url,
            "calculation_method": polygon.calculation_method, "volume_rate": str(polygon.volume_rate or 0),
            "tonnage_rate": str(polygon.tonnage_rate or 0), "waste_types": polygon.waste_types or "",
        }, follow_redirects=False)
        assert rejected.status_code == 400

    trip.status = models.RequestStatus.LOGIST_CONFIRMED
    trip.actual_volume = 20
    trip.actual_tonnage = 8
    trip.sum_driver = 50000
    db.commit()
    page = client.get(f"/polygons?polygon_id={polygon.id}").text
    assert "8" in page and "6 000" in page
    assert "50 000" not in page
    assert 'href="tel:+79995554433"' in page
    assert 'href="https://yandex.ru/maps/?rtext=Шишкино"' in page

    customer_response = client.post("/settings/customers", data={
        "name": "Шишкино — клиент / точка отправления", "address": "Шишкино",
    }, follow_redirects=False)
    assert customer_response.status_code == 302
    assert db.query(models.Customer).filter_by(name="Шишкино — клиент / точка отправления").one()
    db.close()


def test_manual_polygon_cost_is_editable_and_exported_consistently():
    db, admin, logist, driver, vt, vehicle, customer, cargo, polygon, tariff, trip = reset_db()
    polygon.calculation_method = "manual"
    trip.polygon_cost_manual = 4321
    trip.sum_driver = 99999
    db.commit()
    client = client_as(logist)

    edit_page = client.get(f"/requests/{trip.id}/edit")
    assert edit_page.status_code == 200
    assert 'name="polygon_cost_manual"' in edit_page.text

    csv_response = client.get(f"/export/polygon.csv?polygon_id={polygon.id}")
    text = csv_response.content.decode("utf-8-sig")
    assert "Затраты полигона" in text and "4321" in text
    assert "99999" not in text

    summary = client.get("/export/polygons.csv")
    summary_text = summary.content.decode("utf-8-sig")
    assert "Тонны" in summary_text and "4321" in summary_text
    assert "99999" not in summary_text
    db.close()


def test_every_trip_status_has_stable_distinct_ui_class():
    db, admin, *_rest, trip = reset_db()
    client = client_as(admin)
    css = (Path(app_module.root_dir) / "static" / "css" / "app.css").read_text(encoding="utf-8")
    slugs = []
    for status_value in models.RequestStatus:
        slug = app_module.status_slug(status_value)
        slugs.append(slug)
        assert f".status-{slug}" in css
        trip.status = status_value
        db.commit()
        page = client.get(f"/requests/{trip.id}").text
        assert f'class="status status-{slug}"' in page
    assert len(slugs) == len(set(slugs)) == len(models.RequestStatus)
    db.close()


def test_customer_identity_fields_are_editable_in_settings():
    db, admin, *_ = reset_db()
    client = client_as(admin)
    response = client.post("/settings/customers", data={
        "name": "Клиент с реквизитами", "address": "СПб",
        "inn": "7812345678", "bitrix_company_id": "54321",
    }, follow_redirects=False)
    assert response.status_code == 302
    customer = db.query(models.Customer).filter_by(name="Клиент с реквизитами").one()
    assert customer.inn == "7812345678" and customer.bitrix_company_id == 54321
    page = client.get(f"/settings/customers/{customer.id}/edit").text
    assert 'name="inn"' in page and 'value="7812345678"' in page
    assert 'name="bitrix_company_id"' in page and 'value="54321"' in page
    db.close()


def test_driver_day_reports_are_role_scoped_and_validate_dates():
    db, admin, logist, driver, vt, vehicle, *_ = reset_db()
    trip = db.query(models.TripRequest).filter_by(driver_id=driver.id).one()
    trip.planned_date = date(2026, 8, 13)
    other_driver = models.User(
        full_name="Чужой водитель", login="other-day-driver",
        password_hash=app_module.pwd_hash("pass"), role=models.UserRole.DRIVER, is_active=True,
    )
    db.add(other_driver); db.flush()
    foreign_vehicle = models.Vehicle(name="ЧУЖОЙ-АВТО", plate="Х999ХХ78", type_id=vt.id, is_active=True)
    db.add(foreign_vehicle); db.flush()
    foreign = models.DriverDayReport(
        driver_id=other_driver.id, vehicle_id=vehicle.id, report_date=date(2026, 8, 12),
        total_km=999, odometer=99999, fuel_liters=99, comment="ЧУЖОЙ-МАРКЕР",
    )
    db.add(foreign); db.commit()

    driver_client = client_as(driver)
    saved = driver_client.post("/driver/day-report", data={
        "report_date": "2026-08-13", "vehicle_id": str(vehicle.id),
        "total_km": "125.5", "odometer": "45678", "fuel_liters": "42.5",
        "comment": "Свой отчёт",
    }, follow_redirects=False)
    assert saved.status_code == 302
    own = db.query(models.DriverDayReport).filter_by(driver_id=driver.id).one()
    page = driver_client.get("/driver/day-report")
    assert page.status_code == 200 and "Свой отчёт" in page.text
    assert "ЧУЖОЙ-МАРКЕР" not in page.text
    assert "ЧУЖОЙ-АВТО" not in page.text
    forbidden_vehicle = driver_client.post("/driver/day-report", data={
        "report_date": "2026-08-14", "vehicle_id": str(foreign_vehicle.id),
        "total_km": "1", "odometer": "1", "fuel_liters": "1", "comment": "",
    }, follow_redirects=False)
    assert forbidden_vehicle.status_code == 400
    historical_vehicle = driver_client.post("/driver/day-report", data={
        "report_date": "2026-08-14", "vehicle_id": str(vehicle.id),
        "total_km": "1", "odometer": "1", "fuel_liters": "1", "comment": "",
    }, follow_redirects=False)
    assert historical_vehicle.status_code == 400
    assert driver_client.get(f"/driver/day-report/{foreign.id}/edit").status_code == 404
    assert driver_client.get(f"/driver/day-report/{own.id}/edit").status_code == 200
    assert driver_client.get("/reports/day").status_code == 403

    logist_client = client_as(logist)
    report_page = logist_client.get("/reports/day")
    assert report_page.status_code == 200
    assert "Свой отчёт" in report_page.text and "ЧУЖОЙ-МАРКЕР" in report_page.text
    assert logist_client.get("/reports/day?date_from=bad-date").status_code == 400
    assert logist_client.get("/reports/day?date_from=2026-08-14&date_to=2026-08-13").status_code == 400
    db.close()


def test_customer_identity_columns_have_database_unique_indexes():
    db, *_ = reset_db()
    app_module._initialize_database()
    from sqlalchemy import inspect
    indexes = {row["name"]: row for row in inspect(models.engine).get_indexes("customers")}
    assert bool(indexes["uq_customers_inn"]["unique"])
    assert bool(indexes["uq_customers_bitrix_company_id"]["unique"])
    db.close()


def test_admin_edits_user_without_exposing_or_resetting_password_hash():
    db, admin, _, driver, *_ = reset_db()
    old_hash = driver.password_hash
    client = client_as(admin)
    page = client.get(f"/users/{driver.id}/edit")
    assert page.status_code == 200
    assert "Редактирование пользователя" in page.text
    assert old_hash not in page.text
    response = client.post(f"/users/{driver.id}/edit", data={
        "full_name": "Водитель Новый", "login": "edit-driver-new", "password": "",
        "role": "driver", "phone": "+79990000000", "is_active": "on",
    }, follow_redirects=False)
    assert response.status_code == 302
    db.expire_all()
    saved = db.query(models.User).filter_by(id=driver.id).one()
    assert saved.full_name == "Водитель Новый"
    assert saved.login == "edit-driver-new"
    assert saved.password_hash == old_hash
    assert saved.phone == "+79990000000"
    db.close()


def test_settings_records_have_working_edit_routes():
    db, admin, _, _, vt, vehicle, customer, cargo, polygon, tariff, _ = reset_db()
    client = client_as(admin)
    cases = [
        ("vehicles", vehicle.id, {"name": "КАМАЗ новый", "plate": "В222ВВ78", "type_id": str(vt.id), "capacity": "18", "is_active": "on"}, models.Vehicle, "КАМАЗ новый"),
        ("customers", customer.id, {"name": "Заказчик новый", "address": "Новый адрес", "contact": "Иван", "phone": "+7000", "comment": "Важно"}, models.Customer, "Заказчик новый"),
        ("cargo-types", cargo.id, {"name": "Груз новый", "unit": "т", "comment": "Сыпучий"}, models.CargoType, "Груз новый"),
        ("polygons", polygon.id, {"name": "Полигон новый", "address": "Новый полигон", "contact": "Петр", "phone": "+7111", "comment": "Круглосуточно"}, models.Polygon, "Полигон новый"),
        ("tariffs", tariff.id, {"title": "Тариф новый", "kind": "пухтовоз", "vehicle_type_id": str(vt.id), "formula": "trip", "trip_price": "2500", "km_price": "10", "volume_price": "20", "fixed_sum": "100", "is_active": "on"}, models.Tariff, "Тариф новый"),
    ]
    for section, record_id, data, model, expected_name in cases:
        page = client.get(f"/settings/{section}/{record_id}/edit")
        assert page.status_code == 200, section
        assert "Редактирование" in page.text
        response = client.post(f"/settings/{section}/{record_id}/edit", data=data, follow_redirects=False)
        assert response.status_code == 302, section
        db.expire_all()
        row = db.query(model).filter(model.id == record_id).one()
        name = row.title if section == "tariffs" else row.name
        assert name == expected_name
    assert client.get("/settings/vehicles/999999/edit").status_code == 404
    db.close()


def test_vehicle_types_and_routes_are_editable_reference_sections():
    db, admin, _, _, vt, *_ = reset_db()
    route = models.Route(name="Объект старый", load_address="А", unload_address="Б", distance=12)
    db.add(route); db.commit()
    client = client_as(admin)
    settings = client.get("/settings").text
    assert f'/settings/vehicle-types/{vt.id}/edit' in settings
    assert f'/settings/routes/{route.id}/edit' in settings
    assert client.get(f"/settings/vehicle-types/{vt.id}/edit").status_code == 200
    assert client.get(f"/settings/routes/{route.id}/edit").status_code == 200
    assert client.post(f"/settings/vehicle-types/{vt.id}/edit", data={
        "name": "Пухтовоз обновлён", "kind": "пухтовоз", "description": "Описание",
    }, follow_redirects=False).status_code == 302
    assert client.post(f"/settings/routes/{route.id}/edit", data={
        "name": "Объект новый", "load_address": "Новая А", "unload_address": "Новая Б",
        "distance": "25", "comment": "Комментарий",
    }, follow_redirects=False).status_code == 302
    db.expire_all()
    assert db.query(models.VehicleType).filter_by(id=vt.id).one().name == "Пухтовоз обновлён"
    assert db.query(models.Route).filter_by(id=route.id).one().name == "Объект новый"
    db.close()


def test_invalid_role_and_reference_kind_return_400_not_500():
    db, admin, _, driver, vt, *_ = reset_db()
    client = client_as(admin)
    bad_role = client.post(f"/users/{driver.id}/edit", data={
        "full_name": driver.full_name, "login": driver.login, "password": "", "role": "superuser", "phone": "",
    }, follow_redirects=False)
    assert bad_role.status_code == 400
    bad_kind = client.post(f"/settings/vehicle-types/{vt.id}/edit", data={
        "name": vt.name, "kind": "неизвестно", "description": "",
    }, follow_redirects=False)
    assert bad_kind.status_code == 400
    db.close()


def test_route_customer_and_all_tariff_fields_are_editable():
    db, admin, _, _, vt, _, customer, _, _, tariff, _ = reset_db()
    route = models.Route(name="Маршрут тариф", customer_id=None)
    db.add(route); db.commit()
    client = client_as(admin)
    route_saved = client.post(f"/settings/routes/{route.id}/edit", data={
        "name": "Маршрут заказчика", "load_address": "А", "unload_address": "Б",
        "distance": "42.5", "customer_id": str(customer.id), "comment": "Связь",
    }, follow_redirects=False)
    assert route_saved.status_code == 302
    tariff_saved = client.post(f"/settings/tariffs/{tariff.id}/edit", data={
        "title": "Полный тариф", "trip_price": "100", "is_active": "on",
    }, follow_redirects=False)
    assert tariff_saved.status_code == 302
    db.expire_all()
    saved_route = db.query(models.Route).filter_by(id=route.id).one()
    saved_tariff = db.query(models.Tariff).filter_by(id=tariff.id).one()
    assert saved_route.customer_id == customer.id
    assert (saved_tariff.min_km, saved_tariff.max_km, saved_tariff.min_volume, saved_tariff.max_volume) == (0, None, 0, None)
    assert saved_tariff.formula == "trip" and saved_tariff.trip_price == 100
    assert saved_tariff.extra_fee == 0 and saved_tariff.coefficient == 1
    assert saved_tariff.date_from is None and saved_tariff.date_to is None and saved_tariff.comment == ""
    db.close()


def test_edit_rejects_duplicate_unique_reference_values():
    db, admin, _, _, vt, vehicle, _, _, polygon, _, _ = reset_db()
    other_vehicle = models.Vehicle(name="Другая машина", plate="С333СС78", type_id=vt.id)
    other_polygon = models.Polygon(name="Другой полигон")
    db.add_all([other_vehicle, other_polygon]); db.commit()
    client = client_as(admin)
    duplicate_vehicle = client.post(f"/settings/vehicles/{vehicle.id}/edit", data={
        "name": other_vehicle.name, "plate": "Н444НН78", "type_id": str(vt.id), "capacity": "10", "is_active": "on",
    }, follow_redirects=False)
    assert duplicate_vehicle.status_code == 400
    duplicate_polygon = client.post(f"/settings/polygons/{polygon.id}/edit", data={
        "name": other_polygon.name, "address": "Адрес",
    }, follow_redirects=False)
    assert duplicate_polygon.status_code == 400
    db.expire_all()
    assert db.query(models.Vehicle).filter_by(id=vehicle.id).one().name == "КАМАЗ старый"
    assert db.query(models.Polygon).filter_by(id=polygon.id).one().name == "Полигон старый"
    db.close()


def test_driver_does_not_see_or_open_administrative_editing():
    db, admin, logist, driver, _, _, _, _, polygon, _, trip = reset_db()
    driver_client = client_as(driver)
    polygons = driver_client.get("/polygons").text
    assert f"/settings/polygons/{polygon.id}/edit" not in polygons
    assert driver_client.get(f"/settings/polygons/{polygon.id}/edit").status_code == 403
    assert driver_client.get(f"/requests/{trip.id}/edit").status_code == 403
    logist_client = client_as(logist)
    assert logist_client.get(f"/users/{driver.id}/edit").status_code == 403
    assert logist_client.post(f"/settings/polygons/{polygon.id}/delete", follow_redirects=False).status_code == 403
    logist_menu = app_module.menu_for(models.UserRole.LOGIST)
    assert "/settings" not in {item["href"] for item in logist_menu}
    assert "/archive" not in {item["href"] for item in logist_menu}
    assert logist_client.get("/settings").status_code == 403
    assert logist_client.get("/archive").status_code == 403
    db.close()


def test_security_and_integrity_guards_for_editing():
    db, admin, _, driver, vt, vehicle, customer, cargo, polygon, tariff, trip = reset_db()
    customer.name = '<img src=x onerror="alert(1)">'
    trip.status = models.RequestStatus.LOGIST_CONFIRMED
    bad_vt = models.VehicleType(name="Самосвальный тип", kind=models.TripType.SAMOSVAL)
    db.add(bad_vt); db.flush()
    bad_tariff = models.Tariff(title="Чужой тариф", vehicle_type_id=bad_vt.id, kind=models.TripType.SAMOSVAL, is_active=True, trip_price=999)
    db.add(bad_tariff); db.commit()
    client = client_as(admin)

    html = client.get("/settings").text
    assert "<img src=x" not in html and "&lt;img src=x" in html
    assert client.post(f"/users/{driver.id}/edit", headers={"Origin": "https://evil.example"}, data={
        "full_name": driver.full_name, "login": driver.login, "role": "driver",
    }).status_code == 403
    assert client.post(f"/settings/vehicle-types/{vt.id}/delete", follow_redirects=False).status_code == 409
    request_data = {
        "number": trip.number, "planned_date": str(trip.planned_date), "planned_time": "08:00",
        "driver_id": str(driver.id), "vehicle_id": str(vehicle.id), "customer_id": str(customer.id),
        "cargo_type_id": str(cargo.id), "polygon_id": str(polygon.id), "tariff_id": str(bad_tariff.id),
        "load_address": "А", "unload_address": "Б", "route_name": "Р", "km": "1", "volume": "1",
        "trips_count": "1", "kind": "пухтовоз", "comment": "",
    }
    assert client.post(f"/requests/{trip.id}/edit", data=request_data, follow_redirects=False).status_code == 400
    request_data["tariff_id"] = str(tariff.id)
    assert client.post(f"/requests/{trip.id}/edit", data=request_data, follow_redirects=False).status_code == 302
    assert db.query(models.AuditLog).filter_by(section="trip_requests", record_id=trip.id).count() == 1
    self_lock = client.post(f"/users/{admin.id}/edit", data={
        "full_name": admin.full_name, "login": admin.login, "password": "", "role": "logist", "phone": "",
    }, follow_redirects=False)
    assert self_lock.status_code == 400
    old_hash = driver.password_hash
    whitespace = client.post(f"/users/{driver.id}/edit", data={
        "full_name": driver.full_name, "login": driver.login, "password": "   ", "role": "driver", "phone": "",
        "is_active": "on",
    }, follow_redirects=False)
    assert whitespace.status_code == 302
    db.expire_all()
    assert db.query(models.User).filter_by(id=driver.id).one().password_hash == old_hash
    db.close()


def test_request_edit_rejects_incompatible_tariff_and_nonfinite_numbers():
    db, admin, _, driver, _, vehicle, customer, cargo, polygon, _, trip = reset_db()
    wrong_type = models.VehicleType(name="Чужой тип", kind=models.TripType.SAMOSVAL)
    db.add(wrong_type); db.flush()
    wrong_tariff = models.Tariff(title="Несовместимый", vehicle_type_id=wrong_type.id, kind=models.TripType.SAMOSVAL, is_active=True)
    db.add(wrong_tariff); db.commit()
    client = client_as(admin)
    data = {
        "number": trip.number, "planned_date": str(trip.planned_date), "planned_time": "08:00",
        "driver_id": str(driver.id), "vehicle_id": str(vehicle.id), "customer_id": str(customer.id),
        "cargo_type_id": str(cargo.id), "polygon_id": str(polygon.id), "tariff_id": str(wrong_tariff.id),
        "load_address": "А", "unload_address": "Б", "route_name": "Р", "km": "1", "volume": "1",
        "trips_count": "1", "kind": "пухтовоз", "comment": "",
    }
    assert client.post(f"/requests/{trip.id}/edit", data=data, follow_redirects=False).status_code == 400
    data["tariff_id"] = ""
    data["km"] = "NaN"
    assert client.post(f"/requests/{trip.id}/edit", data=data, follow_redirects=False).status_code == 400
    db.close()


def test_tariff_rules_and_formula_are_enforced():
    db, admin, _, driver, vt, vehicle, customer, cargo, polygon, tariff, trip = reset_db()
    client = client_as(admin)
    assert client.post(f"/settings/tariffs/{tariff.id}/edit", data={
        "title": "Проверка", "trip_price": "-1", "is_active": "on",
    }, follow_redirects=False).status_code == 400
    assert client.post(f"/settings/tariffs/{tariff.id}/edit", data={
        "title": "Проверка", "trip_price": "340", "is_active": "on",
    }, follow_redirects=False).status_code == 302
    request_data = {
        "number": trip.number, "planned_date": "2026-08-11", "planned_time": "08:00", "driver_id": str(driver.id),
        "vehicle_id": str(vehicle.id), "customer_id": str(customer.id), "cargo_type_id": str(cargo.id),
        "polygon_id": str(polygon.id), "tariff_id": str(tariff.id), "load_address": "А", "unload_address": "Б",
        "route_name": "Р", "km": "10", "volume": "5", "trips_count": "1", "kind": "пухтовоз", "comment": "",
    }
    assert client.post(f"/requests/{trip.id}/edit", data=request_data, follow_redirects=False).status_code == 302
    db.expire_all()
    assert db.query(models.TripRequest).filter_by(id=trip.id).one().sum_driver == 340
    db.close()


def test_review_security_and_integrity_regressions():
    db, admin, _, driver, vt, vehicle, customer, cargo, polygon, tariff, trip = reset_db()
    app_module.app.dependency_overrides[app_module.get_current_user] = lambda: admin
    no_origin = TestClient(app_module.app)
    assert no_origin.post(f"/users/{driver.id}/edit", data={"full_name": driver.full_name, "login": driver.login, "role": "driver"}).status_code == 403
    client = client_as(admin)
    assert client.post("/users/new", data={"full_name": "Новый", "login": "weak", "password": "   ", "role": "driver"}).status_code == 400
    assert "tariffEl.innerHTML" not in client.get(f"/requests/{trip.id}/edit").text
    tariff.coefficient = 0
    assert app_module._tariff_amount(tariff, 10, 5, 1) == 0
    assert client.post("/settings/routes/999/edit", data={"name": "X"}).status_code == 404
    route = models.Route(name="Маршрут", distance=1)
    db.add(route); db.commit()
    assert client.post(f"/settings/routes/{route.id}/edit", data={"name": "Маршрут", "distance": "-1"}).status_code == 400
    assert client.post(f"/settings/vehicles/{vehicle.id}/edit", data={"name": vehicle.name, "plate": vehicle.plate, "type_id": str(vt.id), "capacity": "-1", "is_active": "on"}).status_code == 400
    assert client.post(f"/settings/vehicle-types/{vt.id}/edit", data={"name": vt.name, "kind": "самосвал"}).status_code == 409
    assert client.post(f"/users/{driver.id}/delete", follow_redirects=False).status_code == 409
    duplicate = models.TripRequest(number="П-ДУБЛЬ", planned_date=date(2026, 8, 11), kind=models.TripType.PUKHTOVOZ)
    db.add(duplicate); db.commit()
    data = {"number": duplicate.number, "planned_date": "2026-08-11", "driver_id": str(driver.id), "vehicle_id": str(vehicle.id), "customer_id": str(customer.id), "cargo_type_id": str(cargo.id), "polygon_id": str(polygon.id), "tariff_id": str(tariff.id), "kind": "пухтовоз"}
    assert client.post(f"/requests/{trip.id}/edit", data=data, follow_redirects=False).status_code == 409
    db.close()


def test_complete_trip_uses_same_tariff_formula():
    db, _, _, driver, _, _, _, _, _, tariff, trip = reset_db()
    tariff.formula, tariff.km_price, tariff.extra_fee, tariff.coefficient = "km", 12, 50, 2
    tariff.min_km, tariff.max_km, tariff.min_volume, tariff.max_volume = 0, 20, 0, 20
    trip.status = models.RequestStatus.IN_WORK
    db.commit()
    client = client_as(driver)
    assert client.post(f"/requests/{trip.id}/complete", data={"actual_km": "10", "actual_volume": "5", "comment": ""}, follow_redirects=False).status_code == 302
    db.expire_all()
    assert db.query(models.TripRequest).filter_by(id=trip.id).one().sum_driver == 340
    db.close()


def test_final_review_tariff_completion_and_salary_guards():
    db, admin, _, driver, vt, vehicle, customer, cargo, polygon, tariff, trip = reset_db()
    client = client_as(admin)
    tariff.coefficient = 0
    tariff.formula = "km"
    tariff.km_price = 12
    tariff.extra_fee = 50
    tariff.min_km, tariff.max_km, tariff.min_volume, tariff.max_volume = 0, 20, 0, 20
    db.commit()
    html = client.get(f"/settings/tariffs/{tariff.id}/edit").text
    assert 'name="title"' in html and 'name="trip_price"' in html
    assert 'name="exact_km"' not in html and 'name="exact_volume"' not in html
    assert 'name="coefficient"' not in html
    tariff.is_active = False
    db.commit()
    edit_data = {"number": trip.number, "planned_date": str(trip.planned_date), "driver_id": str(driver.id), "vehicle_id": str(vehicle.id), "customer_id": str(customer.id), "cargo_type_id": str(cargo.id), "polygon_id": str(polygon.id), "tariff_id": "", "km": "12", "volume": "5", "trips_count": "1", "kind": "пухтовоз"}
    assert client.post(f"/requests/{trip.id}/edit", data=edit_data, follow_redirects=False).status_code == 400
    tariff.is_active = True
    tariff.coefficient = 2
    trip.status = models.RequestStatus.IN_WORK
    db.commit()
    driver_client = client_as(driver)
    assert driver_client.post(f"/requests/{trip.id}/complete", data={"actual_km": "Infinity", "actual_volume": "5"}, follow_redirects=False).status_code == 400
    assert driver_client.post(f"/requests/{trip.id}/complete", data={"actual_km": "0", "actual_volume": "0"}, follow_redirects=False).status_code == 302
    db.expire_all()
    completed = db.query(models.TripRequest).filter_by(id=trip.id).one()
    assert completed.actual_km == 0 and completed.actual_volume == 0 and completed.sum_driver == 100
    assert driver_client.post(f"/requests/{trip.id}/complete", data={"actual_km": "1", "actual_volume": "1"}, follow_redirects=False).status_code == 409
    vehicle.is_active = False
    db.commit()
    db.close()


def test_final_review_vehicle_retype_and_salary_lock():
    db, admin, _, driver, vt, vehicle, customer, cargo, polygon, tariff, trip = reset_db()
    client = client_as(admin)
    other_type = models.VehicleType(name="Другой пухтовоз", kind=models.TripType.PUKHTOVOZ)
    db.add(other_type); db.commit()
    assert client.post(f"/settings/vehicles/{vehicle.id}/edit", data={"name": vehicle.name, "plate": vehicle.plate, "type_id": str(other_type.id), "capacity": "10", "is_active": "on"}).status_code == 409
    trip.sum_driver = tariff.trip_price
    trip.sum_trip = tariff.trip_price
    calc = models.SalaryCalc(driver_id=driver.id, date_from=date(2026, 8, 1), date_to=date(2026, 8, 31), status=models.CalcStatus.DRAFT)
    db.add(calc); db.flush()
    db.add(models.SalaryCalcItem(salary_calc_id=calc.id, trip_request_id=trip.id, sum=trip.sum_driver))
    db.commit()
    data = {"number": trip.number, "planned_date": str(trip.planned_date), "driver_id": str(driver.id), "vehicle_id": str(vehicle.id), "customer_id": str(customer.id), "cargo_type_id": str(cargo.id), "polygon_id": str(polygon.id), "tariff_id": str(tariff.id), "km": "13", "volume": str(trip.volume), "trips_count": str(trip.trips_count), "kind": "пухтовоз", "comment": "изменение"}
    assert client.post(f"/requests/{trip.id}/edit", data=data, follow_redirects=False).status_code == 409
    data["km"] = str(trip.km)
    assert client.post(f"/requests/{trip.id}/edit", data=data, follow_redirects=False).status_code == 409
    db.expire_all()
    assert db.query(models.SalaryCalcItem).filter_by(trip_request_id=trip.id).one().sum == tariff.trip_price
    assert db.query(models.TripRequest).filter_by(id=trip.id).one().comment == "Старый комментарий"
    db.close()


def test_final_requests_and_salary_linked_requests_cannot_be_deleted():
    db, admin, _, _, _, _, _, _, _, _, trip = reset_db()
    client = client_as(admin)
    trip.status = models.RequestStatus.LOGIST_CONFIRMED
    db.commit()
    assert client.post(f"/requests/{trip.id}/delete", follow_redirects=False).status_code == 409
    assert db.query(models.TripRequest).filter_by(id=trip.id).first() is not None
    assert client.post("/requests/999999/delete", follow_redirects=False).status_code == 404
    db.close()


def test_short_edit_password_final_actuals_full_audit_and_salary_button():
    db, admin, _, driver, _, vehicle, customer, cargo, polygon, tariff, trip = reset_db()
    client = client_as(admin)
    assert client.post(f"/users/{driver.id}/edit", data={"full_name": driver.full_name, "login": driver.login, "password": "x", "role": "driver", "is_active": "on"}, follow_redirects=False).status_code == 400
    tariff.formula, tariff.km_price, tariff.extra_fee, tariff.coefficient = "km", 12, 50, 2
    tariff.min_km, tariff.max_km, tariff.min_volume, tariff.max_volume = 0, 20, 0, 20
    trip.status = models.RequestStatus.LOGIST_CONFIRMED
    trip.actual_km, trip.actual_volume = 10, 5
    db.commit()
    data = {"number": trip.number, "planned_date": str(trip.planned_date), "planned_time": "09:30", "driver_id": str(driver.id), "vehicle_id": str(vehicle.id), "customer_id": str(customer.id), "cargo_type_id": str(cargo.id), "polygon_id": str(polygon.id), "tariff_id": str(tariff.id), "load_address": "Новый адрес", "unload_address": "Новая выгрузка", "route_name": "Новый маршрут", "km": "1", "volume": "1", "trips_count": "1", "kind": "пухтовоз", "comment": "Полный аудит"}
    assert client.post(f"/requests/{trip.id}/edit", data=data, follow_redirects=False).status_code == 302
    db.expire_all()
    edited = db.query(models.TripRequest).filter_by(id=trip.id).one()
    assert edited.sum_driver == 340
    audit = db.query(models.AuditLog).filter_by(section="trip_requests", record_id=trip.id).order_by(models.AuditLog.id.desc()).first()
    assert "Новый маршрут" in audit.new_value and "Полный аудит" in audit.new_value and "planned_time" in audit.new_value
    salary = client.get("/salary")
    assert salary.status_code == 200 and f'/requests/{trip.id}/edit' in salary.text
    db.close()


def test_current_security_review_exports_idor_webhook_and_delete_history():
    db, admin, logist, driver, _, vehicle, customer, cargo, polygon, tariff, own = reset_db()
    other = models.User(full_name="Другой водитель", login="other-driver", password_hash=app_module.pwd_hash("secret"), role=models.UserRole.DRIVER, is_active=True)
    db.add(other); db.flush()
    foreign = models.TripRequest(number="П-FOREIGN", planned_date=date.today(), driver_id=other.id, vehicle_id=vehicle.id, customer_id=customer.id, cargo_type_id=cargo.id, polygon_id=polygon.id, tariff_id=tariff.id, kind=models.TripType.PUKHTOVOZ, status=models.RequestStatus.ASSIGNED)
    db.add(foreign); db.flush()
    db.add(models.StatusHistory(trip_request_id=foreign.id, user_id=admin.id, old_status=models.RequestStatus.NEW.value, new_status=models.RequestStatus.ASSIGNED.value)); db.commit()
    app_module.app.dependency_overrides.clear()
    anonymous = TestClient(app_module.app)
    for path in ("/export/report.csv", "/export/report.xlsx", "/export/requests.csv", "/export/salary.xlsx"):
        assert anonymous.get(path).status_code == 401
    driver_client = client_as(driver)
    assert "П-FOREIGN" not in driver_client.get("/requests").text
    assert "П-FOREIGN" not in driver_client.get("/reports").text
    assert driver_client.get(f"/requests/{foreign.id}").status_code == 404
    assert b"P-FOREIGN" not in driver_client.get("/export/requests.csv").content
    integration = models.IntegrationSetting(provider="bitrix24", webhook_url="https://example.invalid", secret="", is_active=True)
    db.add(integration); db.commit()
    assert anonymous.post("/webhook/bitrix24", data={"event": "ONCRMDYNAMICITEMUPDATE"}).status_code == 503
    admin_client = client_as(admin)
    assert admin_client.post(f"/requests/{foreign.id}/delete", follow_redirects=False).status_code == 302
    assert db.query(models.TripRequest).filter_by(id=foreign.id).first() is None
    assert db.query(models.StatusHistory).filter_by(trip_request_id=foreign.id).count() == 0
    db.close()


def test_driver_is_fully_isolated_from_other_driver_html_filters_and_exports():
    db, admin, _, driver, _, vehicle, customer, cargo, polygon, tariff, own = reset_db()
    other = models.User(
        full_name="ВОДИТЕЛЬ-ЧУЖОЙ-МАРКЕР", login="driver-isolation-other",
        password_hash=app_module.pwd_hash("secret"), role=models.UserRole.DRIVER, is_active=True,
    )
    foreign_polygon = models.Polygon(name="ПОЛИГОН-ЧУЖОЙ-МАРКЕР")
    foreign_customer = models.Customer(name="ЗАКАЗЧИК-ЧУЖОЙ-МАРКЕР")
    db.add_all([other, foreign_polygon, foreign_customer]); db.flush()
    own.status = models.RequestStatus.LOGIST_CONFIRMED
    own.sum_driver = 1234
    foreign = models.TripRequest(
        number="РЕЙС-ЧУЖОЙ-МАРКЕР", planned_date=date.today(), driver_id=other.id,
        vehicle_id=vehicle.id, customer_id=foreign_customer.id, cargo_type_id=cargo.id,
        polygon_id=foreign_polygon.id, tariff_id=tariff.id, kind=models.TripType.PUKHTOVOZ,
        status=models.RequestStatus.LOGIST_CONFIRMED, km=8765, volume=7654, sum_driver=987654,
    )
    db.add(foreign); db.commit()
    client = client_as(driver)
    forbidden = (
        "РЕЙС-ЧУЖОЙ-МАРКЕР", "ВОДИТЕЛЬ-ЧУЖОЙ-МАРКЕР",
        "ПОЛИГОН-ЧУЖОЙ-МАРКЕР", "ЗАКАЗЧИК-ЧУЖОЙ-МАРКЕР", "987654",
    )

    html_paths = (
        "/driver", "/requests", "/pukhtovoz", "/samosval",
        "/reports", "/salary", "/polygons",
        f"/reports?driver_id={other.id}", f"/salary?driver_id={other.id}",
    )
    for path in html_paths:
        response = client.get(path)
        assert response.status_code == 200, path
        assert all(marker not in response.text for marker in forbidden), path
    assert client.get(f"/requests/{foreign.id}").status_code == 404
    assert client.get(f"/export/polygon.csv?polygon_id={foreign_polygon.id}").status_code == 404

    csv_paths = (
        f"/export/report.csv?driver_id={other.id}",
        f"/export/requests.csv?driver_id={other.id}",
        f"/export/polygons.csv?driver_id={other.id}",
    )
    for path in csv_paths:
        response = client.get(path)
        assert response.status_code == 200, path
        assert all(marker not in response.text for marker in forbidden), path

    for path in (f"/export/report.xlsx?driver_id={other.id}", f"/export/salary.xlsx?driver_id={other.id}"):
        response = client.get(path)
        assert response.status_code == 200, path
        workbook = load_workbook(io.BytesIO(response.content), read_only=True)
        cells = " ".join("" if value is None else str(value) for row in workbook.active.iter_rows(values_only=True) for value in row)
        assert all(marker not in cells for marker in forbidden), path
    db.close()


def test_current_security_review_audit_zero_lifecycle_and_salary_unique():
    from sqlalchemy.exc import IntegrityError
    db, admin, _, driver, _, vehicle, customer, cargo, polygon, tariff, trip = reset_db()
    client = client_as(admin)
    trip.status = models.RequestStatus.ASSIGNED; trip.actual_km = 0; trip.actual_volume = 0; trip.km = 9; trip.volume = 8
    db.add(models.StatusHistory(trip_request_id=trip.id, user_id=admin.id, old_status=models.RequestStatus.NEW.value, new_status=models.RequestStatus.ASSIGNED.value)); db.commit()
    driver_client = client_as(driver)
    assert driver_client.post(f"/requests/{trip.id}/accept", follow_redirects=False).status_code == 302
    history = db.query(models.StatusHistory).filter_by(trip_request_id=trip.id).order_by(models.StatusHistory.id.desc()).first()
    assert history.old_status == models.RequestStatus.ASSIGNED.value
    assert app_module.calc_sum(trip, tariff) == app_module._tariff_amount(tariff, 0, 0, trip.trips_count)
    client = client_as(admin)
    edit = client.post(f"/settings/customers/{customer.id}/edit", data={"name": "Заказчик аудит", "address": "Адрес", "contact": "Контакт", "phone": "1", "comment": "Текст"}, follow_redirects=False)
    assert edit.status_code == 302
    user_edit = client.post(f"/users/{driver.id}/edit", data={"full_name": "Водитель аудит", "login": driver.login, "password": "", "role": "driver", "is_active": "on"}, follow_redirects=False)
    assert user_edit.status_code == 302
    assert db.query(models.AuditLog).filter_by(section="settings:customers", record_id=customer.id).first()
    assert db.query(models.AuditLog).filter_by(section="users", record_id=driver.id).first()
    calc1 = models.SalaryCalc(driver_id=driver.id, date_from=date.today(), date_to=date.today(), status=models.CalcStatus.DRAFT)
    calc2 = models.SalaryCalc(driver_id=driver.id, date_from=date.today(), date_to=date.today(), status=models.CalcStatus.DRAFT)
    db.add_all([calc1, calc2]); db.flush(); db.add(models.SalaryCalcItem(salary_calc_id=calc1.id, trip_request_id=trip.id, sum=0)); db.commit()
    db.add(models.SalaryCalcItem(salary_calc_id=calc2.id, trip_request_id=trip.id, sum=0))
    try:
        db.commit(); assert False, "duplicate SalaryCalcItem accepted"
    except IntegrityError:
        db.rollback()
    db.close()


def test_edit_buttons_are_visible_in_all_managed_lists():
    db, admin, _, _, _, _, _, _, _, _, trip = reset_db()
    client = client_as(admin)
    trips = client.get("/pukhtovoz").text
    all_requests = client.get("/requests").text
    reports = client.get("/reports").text
    detail = client.get(f"/requests/{trip.id}").text
    users = client.get("/users").text
    settings = client.get("/settings").text
    polygons = client.get("/polygons").text
    assert f'/requests/{trip.id}/edit' in trips
    assert f'/requests/{trip.id}/edit' in all_requests
    assert f'/requests/{trip.id}/edit' in reports
    assert f'/requests/{trip.id}/edit' in detail
    assert "/users/" in users and "/edit" in users and "Редактировать" in users
    for section in ("vehicles", "customers", "cargo-types", "polygons", "tariffs"):
        assert f"/settings/{section}/" in settings
    assert "Редактировать" in settings
    assert "/settings/polygons/" in polygons and "Редактировать" in polygons
    db.close()



def test_final_security_polygon_webhook_admin_trigger_and_inactive_vehicle():
    db, admin, logist, driver, vt, vehicle, customer, cargo, polygon, tariff, own = reset_db()
    other = models.User(full_name="Чужой водитель", login="foreign-driver-final", password_hash=app_module.pwd_hash("secret"), role=models.UserRole.DRIVER, is_active=True)
    db.add(other); db.flush()
    foreign = models.TripRequest(number="П-SECRET-FOREIGN", planned_date=date.today(), driver_id=other.id, vehicle_id=vehicle.id, polygon_id=polygon.id, tariff_id=tariff.id, km=99, volume=99, sum_driver=99999, kind=models.TripType.PUKHTOVOZ, status=models.RequestStatus.ASSIGNED)
    db.add(foreign); db.commit()

    driver_client = client_as(driver)
    detail_csv = driver_client.get(f"/export/polygon.csv?polygon_id={polygon.id}")
    all_csv = driver_client.get("/export/polygons.csv")
    assert detail_csv.status_code == 200 and "П-SECRET-FOREIGN" not in detail_csv.text
    assert all_csv.status_code == 200 and "99999" not in all_csv.text
    page = driver_client.get("/polygons")
    assert page.status_code == 200 and "99999" not in page.text
    assert "Новый полигон" not in page.text

    admin_client = client_as(admin)
    csv_report = admin_client.get("/export/report.csv", params={"q": "П-EDIT"})
    xlsx_report = admin_client.get("/export/report.xlsx", params={"q": "П-EDIT"})
    assert "П-EDIT" in csv_report.text and "П-SECRET-FOREIGN" not in csv_report.text
    workbook = load_workbook(io.BytesIO(xlsx_report.content), read_only=True)
    exported_numbers = [row[0] for row in workbook.active.iter_rows(min_row=2, values_only=True)]
    assert exported_numbers == ["П-EDIT"]
    assert app_module._export_row(["=1+1", "+cmd", "normal"]) == ["'=1+1", "'+cmd", "normal"]
    assert app_module._export_row(["\t=1+1", "\r+cmd", "\n-1", "  @sum"])[0:4] == ["'\t=1+1", "'\r+cmd", "'\n-1", "'  @sum"]
    list_page = admin_client.get("/requests")
    assert f'/requests/{own.id}/delete' in list_page.text
    own.actual_km = 0
    own.actual_volume = 0
    db.commit()
    detail_page = admin_client.get(f"/requests/{own.id}")
    assert "<b>Факт. км:</b> 0" in detail_page.text
    assert "<b>Фактический объём:</b> 0" in detail_page.text

    app_module.app.dependency_overrides.clear()
    app_module.BITRIX_LAST_EVENT = {"received": False}
    response = TestClient(app_module.app).post("/webhook/bitrix24", json={"event": "ONCRMDYNAMICITEMUPDATE"})
    assert response.status_code == 503
    assert app_module.BITRIX_LAST_EVENT == {"received": False}

    vehicle.is_active = False; db.commit()
    form = client_as(admin).get("/pukhtovoz/new")
    assert form.status_code == 200 and vehicle.name not in form.text

    db.delete(admin)
    with pytest.raises(Exception):
        db.commit()
    db.rollback()
    with pytest.raises(Exception):
        models._build_engine("postgresql+definitely_missing_driver://invalid")
    db.close()


def test_startup_requires_secret_key():
    env = os.environ.copy()
    env.pop("SECRET_KEY", None)
    env["DATABASE_URL"] = "sqlite:///:memory:"
    result = subprocess.run(
        [sys.executable, "-c", "import app"],
        cwd=os.path.dirname(os.path.dirname(__file__)),
        env=env,
        capture_output=True,
        text=True,
    )
    assert result.returncode != 0
    assert "SECRET_KEY is required" in result.stderr



def test_bitrix_inbound_rejects_lifecycle_jump_and_salary_locked_changes(monkeypatch):
    db, admin, logist, driver, vt, vehicle, customer, cargo, polygon, tariff, trip = reset_db()
    trip.bitrix_element_id = 777
    trip.bitrix_entity_type_id = 150
    trip.sum_driver = trip.sum_trip = 1000
    settings = models.IntegrationSetting(provider="bitrix24", webhook_url="https://example/rest/1/token/", secret="hook-secret", is_active=True)
    db.add(settings); db.commit()

    def jump_status(item_id, entity_type_id, session, settings=None):
        row = session.query(models.TripRequest).filter_by(id=trip.id).one()
        row.status = models.RequestStatus.LOGIST_CONFIRMED
        return {"ok": True, "action": "update", "trip_id": row.id}
    monkeypatch.setattr(app_module.bitrix, "sync_from_bitrix", jump_status)
    payload = {"event": "ONCRMDYNAMICITEMUPDATE", "data[FIELDS][ID]": "777", "data[FIELDS][ENTITY_TYPE_ID]": "150"}
    client = TestClient(app_module.app)
    response = client.post("/webhook/bitrix24?token=hook-secret", json=payload)
    assert response.status_code == 409
    db.expire_all()
    assert db.query(models.TripRequest).filter_by(id=trip.id).one().status == models.RequestStatus.ASSIGNED

    calc = models.SalaryCalc(driver_id=driver.id, date_from=date.today(), date_to=date.today())
    db.add(calc); db.flush()
    db.add(models.SalaryCalcItem(salary_calc_id=calc.id, trip_request_id=trip.id, sum=trip.sum_driver or 0)); db.commit()
    def change_salary_value(item_id, entity_type_id, session, settings=None):
        row = session.query(models.TripRequest).filter_by(id=trip.id).one()
        row.actual_volume = 0
        return {"ok": True, "action": "update", "trip_id": row.id}
    monkeypatch.setattr(app_module.bitrix, "sync_from_bitrix", change_salary_value)
    response = client.post("/webhook/bitrix24?token=hook-secret", json=payload)
    assert response.status_code == 409
    db.expire_all()
    assert db.query(models.TripRequest).filter_by(id=trip.id).one().actual_volume is None

    def change_actual_tonnage(item_id, entity_type_id, session, settings=None):
        row = session.query(models.TripRequest).filter_by(id=trip.id).one()
        row.actual_tonnage = 0
        return {"ok": True, "action": "update", "trip_id": row.id}
    monkeypatch.setattr(app_module.bitrix, "sync_from_bitrix", change_actual_tonnage)
    response = client.post("/webhook/bitrix24?token=hook-secret", json=payload)
    assert response.status_code == 409
    db.expire_all()
    assert db.query(models.TripRequest).filter_by(id=trip.id).one().actual_tonnage is None

    def change_operational_values(item_id, entity_type_id, session, settings=None):
        row = session.query(models.TripRequest).filter_by(id=trip.id).one()
        row.planned_time = "23:59"
        row.site_contact_phone = "+79990000000"
        row.polygon_cost_manual = 1234
        return {"ok": True, "action": "update", "trip_id": row.id}
    monkeypatch.setattr(app_module.bitrix, "sync_from_bitrix", change_operational_values)
    response = client.post("/webhook/bitrix24?token=hook-secret", json=payload)
    assert response.status_code == 409
    db.expire_all()
    locked = db.query(models.TripRequest).filter_by(id=trip.id).one()
    assert locked.planned_time == "08:00"
    assert locked.site_contact_phone is None
    assert locked.polygon_cost_manual is None

    db.query(models.SalaryCalcItem).delete(); db.query(models.SalaryCalc).delete()
    trip.status = models.RequestStatus.LOGIST_CONFIRMED
    trip.actual_volume = None
    db.commit()
    response = client.post("/webhook/bitrix24?token=hook-secret", json=payload)
    assert response.status_code == 409
    db.expire_all()
    assert db.query(models.TripRequest).filter_by(id=trip.id).one().actual_volume is None
    db.close()


def test_delete_request_removes_attachment_rows_and_files(tmp_path, monkeypatch):
    db, admin, _, _, _, _, _, _, _, _, trip = reset_db()
    stored = tmp_path / "attachment.pdf"
    stored.write_bytes(b"%PDF-test")
    db.add(models.Attachment(
        trip_request_id=trip.id, filename="attachment.pdf", content_type="application/pdf",
        size=9, path=str(stored), content=b"%PDF-test",
    ))
    db.commit()
    monkeypatch.setattr(app_module.bitrix, "delete_trip", lambda req, session: {"skipped": True})
    response = client_as(admin).post(f"/requests/{trip.id}/delete", follow_redirects=False)
    assert response.status_code == 302
    assert db.query(models.TripRequest).filter_by(id=trip.id).first() is None
    assert db.query(models.Attachment).filter_by(trip_request_id=trip.id).count() == 0
    assert not stored.exists()
    db.close()


def test_salary_locked_request_rejects_all_operational_edits():
    db, admin, _, driver, _, vehicle, customer, cargo, polygon, tariff, trip = reset_db()
    trip.sum_driver = trip.sum_trip = 1000
    calc = models.SalaryCalc(driver_id=driver.id, date_from=date.today(), date_to=date.today())
    db.add(calc); db.flush()
    db.add(models.SalaryCalcItem(salary_calc_id=calc.id, trip_request_id=trip.id, sum=trip.sum_driver or 0))
    db.commit()
    response = client_as(admin).post(f"/requests/{trip.id}/edit", data={
        "number": trip.number, "planned_date": str(trip.planned_date), "planned_time": "23:59",
        "driver_id": str(driver.id), "vehicle_id": str(vehicle.id),
        "customer_id": str(customer.id), "cargo_type_id": str(cargo.id),
        "polygon_id": str(polygon.id), "tariff_id": str(tariff.id),
        "load_address": trip.load_address, "unload_address": trip.unload_address,
        "route_name": trip.route_name, "km": str(trip.km), "volume": str(trip.volume),
        "tonnage": "12", "trips_count": str(trip.trips_count), "kind": trip.kind.value,
        "comment": "Изменён", "polygon_cost_manual": "1234",
        "site_contact_name": "Другой", "site_contact_phone": "+79990000000",
        "site_contact_comment": "Изменён",
    }, follow_redirects=False)
    assert response.status_code == 409
    db.expire_all()
    locked = db.query(models.TripRequest).filter_by(id=trip.id).one()
    assert locked.planned_time == "08:00"
    assert locked.tonnage is None
    assert locked.polygon_cost_manual is None
    assert locked.comment == "Старый комментарий"
    db.close()


def test_sqlite_attachment_limit_reserves_write_lock_before_count(monkeypatch):
    db, *_ = reset_db()
    statements = []
    original_execute = db.execute

    def recording_execute(statement, *args, **kwargs):
        statements.append(str(statement))
        return original_execute(statement, *args, **kwargs)

    monkeypatch.setattr(db, "execute", recording_execute)
    app_module._lock_attachment_parent(db, 1)
    assert statements and statements[0].upper() == "BEGIN IMMEDIATE"
    db.rollback()
    db.close()


def test_empty_reports_row_spans_every_column():
    db, admin, _, driver, *_ = reset_db()
    db.query(models.TripRequest).delete(); db.commit()
    assert 'colspan="10"' in client_as(admin).get("/reports").text
    assert 'colspan="9"' in client_as(driver).get("/reports").text
    db.close()


def test_bitrix_inbound_create_writes_initial_status_history(monkeypatch):
    db, admin, logist, driver, vt, vehicle, customer, cargo, polygon, tariff, _ = reset_db()
    settings = models.IntegrationSetting(provider="bitrix24", webhook_url="https://example/rest/1/token/", secret="hook-secret", is_active=True)
    db.add(settings); db.commit()

    def create_trip(item_id, entity_type_id, session, settings=None):
        trip = models.TripRequest(
            number="П-BITRIX-NEW", planned_date=date.today(), kind=models.TripType.PUKHTOVOZ,
            status=models.RequestStatus.NEW, bitrix_element_id=item_id, bitrix_entity_type_id=entity_type_id,
        )
        session.add(trip); session.flush()
        return {"ok": True, "action": "add", "trip_id": trip.id}

    monkeypatch.setattr(app_module.bitrix, "sync_from_bitrix", create_trip)
    payload = {"event": "ONCRMDYNAMICITEMADD", "data[FIELDS][ID]": "778", "data[FIELDS][ENTITY_TYPE_ID]": "150"}
    response = TestClient(app_module.app).post("/webhook/bitrix24?token=hook-secret", json=payload)
    assert response.status_code == 200
    trip_id = response.json()["trip_id"]
    history = db.query(models.StatusHistory).filter_by(trip_request_id=trip_id).all()
    assert len(history) == 1
    assert history[0].old_status is None
    assert history[0].new_status == models.RequestStatus.NEW.value
    db.close()


def test_logout_is_post_only_and_archived_vehicle_cannot_be_retyped():
    db, admin, logist, driver, vt, vehicle, customer, cargo, polygon, tariff, trip = reset_db()
    client = client_as(admin)
    assert client.get("/logout", follow_redirects=False).status_code == 405
    logout_response = client.post("/logout", follow_redirects=False)
    assert logout_response.status_code == 302
    assert logout_response.headers["location"] == "/login"

    archived_vehicle = models.Vehicle(name="Архивная машина", plate="Х999ХХ78", type_id=vt.id, is_active=True)
    other_type = models.VehicleType(name="Другой архивный тип", kind=models.TripType.SAMOSVAL)
    db.add_all([archived_vehicle, other_type]); db.flush()
    db.add(models.TripArchive(
        origin_id=999, number="АРХ-999", planned_date=date.today(), vehicle_id=archived_vehicle.id,
        kind=models.TripType.PUKHTOVOZ, status=models.RequestStatus.LOGIST_CONFIRMED.value,
    ))
    db.commit()
    response = client.post(f"/settings/vehicles/{archived_vehicle.id}/edit", data={
        "name": archived_vehicle.name, "plate": archived_vehicle.plate,
        "type_id": str(other_type.id), "capacity": "10", "is_active": "on",
    }, follow_redirects=False)
    assert response.status_code == 409
    db.expire_all()
    assert db.query(models.Vehicle).filter_by(id=archived_vehicle.id).one().type_id == vt.id
    db.close()


def test_bitrix_number_match_cannot_bypass_salary_lock(monkeypatch):
    db, admin, logist, driver, vt, vehicle, customer, cargo, polygon, tariff, trip = reset_db()
    trip.bitrix_element_id = None
    trip.bitrix_entity_type_id = None
    trip.status = models.RequestStatus.LOGIST_CONFIRMED
    trip.actual_volume = None
    calc = models.SalaryCalc(driver_id=driver.id, date_from=date.today(), date_to=date.today(), status=models.CalcStatus.DRAFT)
    db.add(calc); db.flush()
    db.add(models.SalaryCalcItem(salary_calc_id=calc.id, trip_request_id=trip.id, sum=trip.sum_driver or 0))
    integration = models.IntegrationSetting(provider="bitrix24", webhook_url="https://example.invalid/rest/", secret="number-match-secret", is_active=True)
    db.add(integration); db.commit()

    monkeypatch.setattr(app_module.bitrix, "extract_event_identifiers", lambda payload: ("ONCRMDYNAMICITEMUPDATE", 987, 177))
    calls = {"count": 0}
    def sync_by_number(item_id, entity_type_id, session, settings=None):
        calls["count"] += 1
        matched = session.query(models.TripRequest).filter_by(number=trip.number, kind=trip.kind).one()
        matched.actual_volume = 0
        matched.bitrix_element_id = item_id
        matched.bitrix_entity_type_id = entity_type_id
        session.flush()
        return {"ok": True, "action": "update", "trip_id": matched.id}
    monkeypatch.setattr(app_module.bitrix, "sync_from_bitrix", sync_by_number)

    response = client_as(admin).post("/webhook/bitrix24?token=number-match-secret", json={"event": "ONCRMDYNAMICITEMUPDATE"})
    assert response.status_code == 409
    assert calls["count"] == 1
    db.expire_all()
    saved = db.query(models.TripRequest).filter_by(id=trip.id).one()
    assert saved.actual_volume is None
    assert saved.bitrix_element_id is None
    db.close()


def test_integration_secrets_stay_out_of_dom_blank_update_preserves_and_xlsx_is_in_memory(monkeypatch):
    db, admin, logist, driver, vt, vehicle, customer, cargo, polygon, tariff, trip = reset_db()
    integration = models.IntegrationSetting(
        provider="bitrix24", webhook_url="https://secret.example/rest/1/private/",
        secret="never-render-this-secret", responsible_id="7", is_active=True,
    )
    db.add(integration); db.commit()
    client = client_as(admin)
    settings_page = client.get("/settings")
    assert settings_page.status_code == 200
    assert "https://secret.example/rest/1/private/" not in settings_page.text
    assert "never-render-this-secret" not in settings_page.text
    assert 'name="secret" value=""' in settings_page.text
    assert 'name="webhook_url" value=""' in settings_page.text
    monkeypatch.setattr(
        app_module.bitrix, "find_smart_process_ids",
        lambda url: {"_error": f"connection failed for {url}"},
    )
    test_page = client.get("/settings/bitrix/test")
    assert test_page.status_code == 200
    assert "https://secret.example/rest/1/private/" not in test_page.text
    assert "never-render-this-secret" not in test_page.text
    assert "connection failed" not in test_page.text
    assert "Webhook:</b> настроен" in test_page.text

    response = client.post("/settings/integrations", data={
        "provider": "bitrix24", "webhook_url": "", "secret": "", "responsible_id": "8",
    }, follow_redirects=False)
    assert response.status_code == 302
    db.expire_all()
    saved = db.query(models.IntegrationSetting).filter_by(provider="bitrix24").one()
    assert saved.webhook_url == "https://secret.example/rest/1/private/"
    assert saved.secret == "never-render-this-secret"
    assert saved.responsible_id == "8" and saved.is_active is True

    report = client.get("/export/report.xlsx")
    salary = client.get("/export/salary.xlsx")
    expected_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert report.status_code == 200 and report.headers["content-type"] == expected_type
    assert salary.status_code == 200 and salary.headers["content-type"] == expected_type
    assert load_workbook(io.BytesIO(report.content), read_only=True).active.max_row >= 1
    assert load_workbook(io.BytesIO(salary.content), read_only=True).active.max_row >= 1
    db.close()


def test_render_binds_health_before_slow_database_initialization(tmp_path):
    import socket
    import sqlite3
    import time
    import urllib.error
    import urllib.request

    db_path = tmp_path / "render-delayed.db"
    lock = sqlite3.connect(db_path)
    lock.execute("BEGIN EXCLUSIVE")
    with socket.socket() as probe:
        probe.bind(("127.0.0.1", 0))
        port = probe.getsockname()[1]
    env = os.environ.copy()
    env.pop("RENDER", None)
    env.pop("RENDER_SERVICE_ID", None)
    env.update({
        "DATABASE_URL": f"sqlite:///{db_path.as_posix()}",
        "SECRET_KEY": "render-delayed-start-test",
        "PYTHONPATH": str(Path(app_module.root_dir)),
        "PORT": str(port),
    })
    proc = subprocess.Popen(
        [sys.executable, "-m", "uvicorn", "app:app", "--host", "127.0.0.1", "--port", str(port), "--workers", "1"],
        cwd=Path(app_module.root_dir), env=env,
        stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True,
    )
    try:
        health = None
        health_status = None
        for _ in range(80):
            if proc.poll() is not None:
                raise AssertionError(proc.stdout.read())
            try:
                with urllib.request.urlopen(f"http://127.0.0.1:{port}/healthz", timeout=1) as response:
                    health_status = response.status
                    health = response.read().decode()
                break
            except urllib.error.HTTPError as exc:
                if exc.code == 503:
                    health_status = exc.code
                    health = exc.read().decode()
                    break
            except Exception:
                pass
            time.sleep(0.1)
        assert health_status == 503
        assert health and '"status":"starting"' in health
        with pytest.raises(urllib.error.HTTPError) as blocked:
            urllib.request.urlopen(f"http://127.0.0.1:{port}/reports", timeout=2)
        assert blocked.value.code == 503

        lock.rollback()
        lock.close()
        for _ in range(100):
            try:
                with urllib.request.urlopen(f"http://127.0.0.1:{port}/healthz", timeout=1) as response:
                    if '"status":"ready"' in response.read().decode():
                        break
            except Exception:
                pass
            time.sleep(0.1)
        else:
            raise AssertionError("Database initialization did not become ready")
    finally:
        try:
            lock.close()
        except Exception:
            pass
        if proc.poll() is None:
            proc.terminate()
            try:
                proc.wait(timeout=10)
            except subprocess.TimeoutExpired:
                proc.kill()


def test_empty_database_has_no_demo_accounts_and_secure_bootstrap_is_explicit(tmp_path):
    ROOT = Path(app_module.root_dir)
    source = (ROOT / "app.py").read_text(encoding="utf-8")
    for predictable in ("admin123", "logist123", "driver123"):
        assert predictable not in source

    code = "import app; from backend.models import SessionLocal,User; db=SessionLocal(); print('USERS='+str(db.query(User).count())); db.close()"
    base_env = os.environ.copy()
    base_env.update({"SECRET_KEY": "subprocess-test-key", "PYTHONPATH": str(ROOT)})
    base_env.pop("BOOTSTRAP_ADMIN_PASSWORD", None)
    empty_db = tmp_path / "empty.db"
    no_bootstrap = subprocess.run(
        [sys.executable, "-c", code], cwd=ROOT,
        env={**base_env, "DATABASE_URL": f"sqlite:///{empty_db.as_posix()}"},
        capture_output=True, text=True, timeout=60,
    )
    assert no_bootstrap.returncode == 0 and "USERS=0" in no_bootstrap.stdout

    secure_db = tmp_path / "secure.db"
    secure_bootstrap = subprocess.run(
        [sys.executable, "-c", code], cwd=ROOT,
        env={**base_env, "DATABASE_URL": f"sqlite:///{secure_db.as_posix()}", "BOOTSTRAP_ADMIN_PASSWORD": "strong-bootstrap-2026"},
        capture_output=True, text=True, timeout=60,
    )
    assert secure_bootstrap.returncode == 0 and "USERS=1" in secure_bootstrap.stdout


def test_bitrix_diagnostics_never_expose_urls_tokens_or_raw_errors(monkeypatch, capsys):
    db, admin, logist, driver, vt, vehicle, customer, cargo, polygon, tariff, trip = reset_db()
    integration = models.IntegrationSetting(
        provider="bitrix24", webhook_url="https://secret.example/rest/1/private-token/",
        secret="private-application-token", is_active=True,
    )
    db.add(integration); db.commit()
    leaked = "https://secret.example/rest/1/private-token/?token=private-application-token"
    monkeypatch.setattr(app_module.bitrix, "extract_event_identifiers", lambda payload: ("ONCRMDYNAMICITEMUPDATE", 42, 177))
    monkeypatch.setattr(app_module.bitrix, "sync_from_bitrix", lambda *args, **kwargs: {"error": f"request failed: {leaked}"})
    app_module.BITRIX_LAST_EVENT = {}
    client = client_as(admin)

    webhook = client.post("/webhook/bitrix24?token=private-application-token", json={"event": "ONCRMDYNAMICITEMUPDATE"})
    assert webhook.status_code == 400
    assert webhook.json() == {"error": "bitrix_sync_error"}
    assert leaked not in webhook.text and "private-application-token" not in webhook.text

    app_module.BITRIX_LAST_OUTBOUND = {
        "attempted": True, "request_id": trip.id,
        "result": {"error": f"outbound failed: {leaked}", "url": leaked},
        "webhook_url": leaked,
    }
    status_response = client.get("/settings/bitrix/status")
    assert status_response.status_code == 200
    assert leaked not in status_response.text
    assert "private-application-token" not in status_response.text
    payload = status_response.json()
    assert payload["inbound"]["result"] == {"error": "bitrix_sync_error"}
    assert payload["outbound"]["result"] == {"error": "bitrix_sync_error"}

    def raise_with_secret(*args, **kwargs):
        raise RuntimeError(leaked)

    monkeypatch.setattr(app_module.bitrix, "sync_from_bitrix", raise_with_secret)
    exception_response = client.post(
        "/webhook/bitrix24?token=private-application-token",
        json={"event": "ONCRMDYNAMICITEMUPDATE"},
    )
    assert exception_response.status_code == 400
    assert exception_response.json() == {"error": "bitrix_sync_error"}

    monkeypatch.setattr(app_module.bitrix, "resolve_process_entity", lambda url, kind: "1088")
    monkeypatch.setattr(app_module.bitrix, "get_element_fields", lambda url, entity: {"title": {"title": "Название"}})
    monkeypatch.setattr(app_module.bitrix, "_http_post", lambda *args, **kwargs: {"error": leaked})
    provider_result = app_module.bitrix.sync_trip(trip, db, settings=integration)
    assert provider_result["error"] == leaked
    captured = capsys.readouterr()
    assert leaked not in captured.out and leaked not in captured.err
    assert "private-application-token" not in captured.out and "private-application-token" not in captured.err
    db.close()
